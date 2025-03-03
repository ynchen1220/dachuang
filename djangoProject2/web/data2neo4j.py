# coding: utf-8
from py2neo import Graph, Node, Relationship
from openpyxl import load_workbook

if __name__ == '__main__':
    # 连接neo4j数据库，输入地址、用户名、密码
    # graph = Graph('http://localhost:7474', username='neo4j', password='123456')
    graph = Graph("http://localhost:7474", auth=("neo4j", "jiayi031220"))
    workbook = load_workbook(u'../all.xlsx')
    booksheet = workbook.active
    # 获取sheet页的行数据
    # rows = booksheet.rows
    # 获取sheet页的列数据
    # columns = booksheet.columns
    i = 1
    graph.run('match (n) detach delete n')  # 删除所有节点及其关系
    loop = True
    print("neo4j已连接")
while loop:
    s = []
    i += 1
    if i > 218488: #取前1000个行
        break
    # line = [col.value for col in row]
    entity_1 = booksheet.cell(row=i, column=1).value
    if entity_1 == None:
        loop = False
        break
    relation_data = booksheet.cell(row=i, column=2).value
    entity_2 = booksheet.cell(row=i, column=3).value
    # print(entity_1, relation_data, entity_2)

    entity_node_1 = Node(entity_1, label='patent', name=entity_1)
    s.append(entity_node_1)

    entity_node_2 = Node(entity_2, label='patent', name=entity_2)
    s.append(entity_node_2)

    relationship = Relationship(entity_node_1, relation_data, entity_node_2)
    graph.merge(entity_node_1, entity_1, "name")
    graph.merge(entity_node_2, entity_2, "name")
    graph.create(relationship)
print("数据已导入neo4j")